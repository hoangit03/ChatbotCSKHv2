"""
scripts/debug_qa.py

Script debug Q&A pipeline — chạy độc lập, không cần uvicorn.

Kiểm tra từng bước:
  1. Đọc file Excel trực tiếp → hiện cấu trúc
  2. Load vào QAStore qua ImportQAUseCase
  3. Test matching với câu hỏi mẫu → in điểm chi tiết
  4. Tìm top-N match cho câu hỏi tùy ý

Chạy:
    cd d:\\CT-Group\\ChatbotCSKH_BH
    .venv\\Scripts\\activate
    python scripts/debug_qa.py

Hoặc với câu hỏi tùy ý:
    python scripts/debug_qa.py "Chủ đầu tư là ai?"
"""
from __future__ import annotations

import os
import sys
import re
import unicodedata
from pathlib import Path

# Thêm project root vào sys.path
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# ── Màu ANSI ──────────────────────────────────────────────────────
GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
CYAN   = "\033[96m"
BOLD   = "\033[1m"
RESET  = "\033[0m"

def ok(msg):   print(f"{GREEN}✓ {msg}{RESET}")
def warn(msg): print(f"{YELLOW}⚠ {msg}{RESET}")
def err(msg):  print(f"{RED}✗ {msg}{RESET}")
def info(msg): print(f"{CYAN}  {msg}{RESET}")
def header(msg): print(f"\n{BOLD}{'─'*60}\n  {msg}\n{'─'*60}{RESET}")


# ═══════════════════════════════════════════════════════════════════
# BƯỚC 1 — Đọc file Excel trực tiếp
# ═══════════════════════════════════════════════════════════════════

def step1_read_excel(excel_path: Path):
    header("BƯỚC 1: Đọc file Excel trực tiếp")

    if not excel_path.exists():
        err(f"File không tồn tại: {excel_path}")
        err("Kiểm tra lại QA_AUTOLOAD_FILE trong .env")
        return None, None

    ok(f"File tồn tại: {excel_path}")

    try:
        import openpyxl
    except ImportError:
        err("Thiếu openpyxl. Chạy: pip install openpyxl")
        return None, None

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb.active
    info(f"Sheet: {ws.title}")
    info(f"Tổng số dòng: {ws.max_row}  |  Cột: {ws.max_column}")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        err("File trống!")
        return None, None

    # Header
    header_row = rows[0]
    print(f"\n  Header (dòng 1): {header_row}")

    # 5 dòng đầu
    print(f"\n  5 dòng dữ liệu đầu tiên:")
    for i, row in enumerate(rows[1:6], start=2):
        # Truncate dài
        truncated = tuple(
            (str(v)[:60] + "…" if v and len(str(v)) > 60 else v)
            for v in row
        )
        print(f"    Row {i}: {truncated}")

    # Đếm dòng có dữ liệu hợp lệ (không rỗng ở cột 1 và 2)
    valid = 0
    empty_answer = 0
    empty_question = 0
    for row in rows[1:]:
        q = str(row[1] or "").strip() if len(row) > 1 else ""
        a = str(row[2] or "").strip() if len(row) > 2 else ""
        if q and a:
            valid += 1
        elif q and not a:
            empty_answer += 1
        elif not q:
            empty_question += 1

    ok(f"Dòng hợp lệ (có cả question + answer): {valid}")
    if empty_answer:
        warn(f"Dòng có question nhưng thiếu answer: {empty_answer} (sẽ bị skip)")
    if empty_question:
        info(f"Dòng trống / header nhóm: {empty_question} (bình thường)")

    return rows, header_row


# ═══════════════════════════════════════════════════════════════════
# BƯỚC 2 — Load vào QAStore
# ═══════════════════════════════════════════════════════════════════

def step2_load_store(excel_path: Path, project_name: str):
    header("BƯỚC 2: Load dữ liệu vào QAStore")

    # Setup minimal settings (không load .env để tránh dependency)
    os.environ.setdefault("APP_SECRET_KEY", "debug-secret-key-for-testing-only-32ch")
    os.environ.setdefault("LLM_PROVIDER", "openai")
    os.environ.setdefault("OPENAI_API_KEY", "sk-debug-key-not-used")

    try:
        from app.agent.tools.qa_tool import QAStore
        from app.application.usecases.import_qa import ImportQAUseCase, ImportQARequest
    except Exception as e:
        err(f"Import lỗi: {e}")
        err("Chạy từ thư mục root của project: cd d:\\CT-Group\\ChatbotCSKH_BH")
        return None

    store = QAStore()
    uc = ImportQAUseCase(qa_store=store)

    file_bytes = excel_path.read_bytes()
    try:
        result = uc.execute(ImportQARequest(
            file_bytes=file_bytes,
            file_name=excel_path.name,
            project_name=project_name,
        ))
    except Exception as e:
        err(f"ImportQA lỗi: {e}")
        import traceback; traceback.print_exc()
        return None

    ok(f"Đã import: {result.imported} Q&A items")
    if result.skipped:
        info(f"Skipped: {result.skipped} dòng (header nhóm / thiếu data)")
    if result.errors:
        warn(f"Có {len(result.errors)} lỗi:")
        for e in result.errors[:10]:
            print(f"    {RED}{e}{RESET}")

    # Hiện 3 Q&A đầu tiên trong store
    items = list(store._items.values())
    print(f"\n  Mẫu 3 Q&A đầu trong store:")
    for item in items[:3]:
        print(f"    [{item.project_name}] Q: {item.question[:70]}")
        print(f"           A: {item.answer[:70]}")
        if item.keywords:
            print(f"           KW: {item.keywords}")
        print()

    return store


# ═══════════════════════════════════════════════════════════════════
# BƯỚC 3 — Debug matching với điểm số chi tiết
# ═══════════════════════════════════════════════════════════════════

def step3_debug_match(store, query: str, project_name: str | None = None, top_k: int = 5):
    header(f"BƯỚC 3: Debug matching cho query")
    print(f'  Query: "{BOLD}{query}{RESET}"')
    if project_name:
        print(f'  Project filter: "{project_name}"')

    from app.agent.tools.qa_tool import (
        _normalize, _tokenize, _bigrams, _score, _ABBREVIATIONS
    )

    q_norm = _normalize(query)
    q_tokens = set(_tokenize(q_norm))
    q_bigrams = set(_bigrams(q_tokens))

    print(f"\n  Sau normalize: {CYAN}{q_norm!r}{RESET}")
    print(f"  Tokens (sau bỏ stop words): {CYAN}{sorted(q_tokens)}{RESET}")
    print(f"  Bigrams: {CYAN}{sorted(q_bigrams)}{RESET}")

    # Tính score với TẤT CẢ items
    scored = []
    items = list(store._items.values())
    for item in items:
        if not item.is_active:
            continue
        if project_name and item.project_name != project_name:
            continue
        score = _score(q_norm, q_tokens, q_bigrams, item)
        scored.append((score, item))

    scored.sort(key=lambda x: x[0], reverse=True)

    print(f"\n  Tổng số Q&A items đã scan: {len(scored)}")
    print(f"\n  Top {top_k} kết quả (kể cả dưới threshold):\n")

    for rank, (score, item) in enumerate(scored[:top_k], 1):
        item_norm = _normalize(item.question)
        item_tokens = set(_tokenize(item_norm))
        item_bigrams = set(_bigrams(item_tokens))

        # Tính breakdown
        from app.agent.tools.qa_tool import _jaccard
        unigram = _jaccard(q_tokens, item_tokens)
        bigram  = _jaccard(q_bigrams, item_bigrams) if q_bigrams and item_bigrams else 0.0

        kw_score = 0.0
        if item.keywords:
            kw_norms = [_normalize(kw) for kw in item.keywords]
            hits = sum(1 for kw in kw_norms if kw and kw in q_norm)
            kw_score = min(0.4 * hits, 0.8)

        sub_bonus = 0.25 if len(q_norm) > 5 and (q_norm in item_norm or item_norm in q_norm) else 0.0
        exact = 1.0 if q_norm == item_norm else 0.0

        # Màu theo score
        col = GREEN if score >= 0.30 else (YELLOW if score >= 0.15 else RED)

        print(f"  #{rank}  Score: {col}{score:.3f}{RESET}")
        print(f"       Q: {item.question}")
        print(f"       Score breakdown:")
        if exact: print(f"         exact_match  = {GREEN}1.000{RESET} ← EXACT!")
        print(f"         unigram_jaccard = {unigram:.3f} × 0.5 = {unigram*0.5:.3f}")
        print(f"         bigram_jaccard  = {bigram:.3f} × 0.3 = {bigram*0.3:.3f}")
        print(f"         kw_score        = {kw_score:.3f}")
        print(f"         sub_bonus       = {sub_bonus:.3f}")

        # Hiển thị token overlap
        common_tokens = q_tokens & item_tokens
        if common_tokens:
            print(f"         tokens chung: {CYAN}{sorted(common_tokens)}{RESET}")
        else:
            print(f"         {RED}KHÔNG có token chung!{RESET}")
            print(f"         Item tokens: {sorted(item_tokens)}")
        print()


# ═══════════════════════════════════════════════════════════════════
# BƯỚC 4 — Test với nhiều câu hỏi mẫu từ file Q&A
# ═══════════════════════════════════════════════════════════════════

def step4_batch_test(store, project_name: str | None):
    header("BƯỚC 4: Batch test với câu hỏi mẫu từ store")

    from app.agent.tools.qa_tool import _normalize, _tokenize, _bigrams, _score

    items = [i for i in store._items.values() if i.is_active]
    if not items:
        warn("Store rỗng!")
        return

    # Lấy 10 Q&A đầu tiên và test với đúng câu hỏi của chúng
    print(f"  Test exact match với 10 câu hỏi đầu tiên từ store:\n")
    passed = 0
    failed = 0

    for item in items[:10]:
        exact_query = item.question
        q_norm = _normalize(exact_query)
        q_tokens = set(_tokenize(q_norm))
        q_bigrams = set(_bigrams(q_tokens))

        score = _score(q_norm, q_tokens, q_bigrams, item)

        if score >= 0.30:
            print(f"  {GREEN}✓ PASS{RESET} ({score:.3f}) Q: {item.question[:60]}")
            passed += 1
        else:
            print(f"  {RED}✗ FAIL{RESET} ({score:.3f}) Q: {item.question[:60]}")
            failed += 1

    print(f"\n  Kết quả: {GREEN}{passed} PASS{RESET} / {RED}{failed} FAIL{RESET} trong {min(10, len(items))} câu")

    if failed > 0:
        warn("Có câu fail → vấn đề nằm ở thuật toán scoring hoặc normalization")
    else:
        ok("Tất cả exact match đều pass → matching algorithm OK")
        info("Nếu hỏi câu tương tự (không copy y hệt) mà fail → cần improve semantic matching")


# ═══════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════

def main():
    # ── Config ──
    EXCEL_PATH   = ROOT / "data" / "Q&A.xlsx"
    PROJECT_NAME = "EmeraldRiverside"

    # Câu hỏi từ command line hoặc dùng mặc định
    if len(sys.argv) > 1:
        test_queries = [" ".join(sys.argv[1:])]
    else:
        # Các câu hỏi mẫu để test
        test_queries = [
            "Dự án này do CĐT tự thực hiện hay liên doanh/liên kết?",
            "Ai là người đại diện pháp luật của Công ty?",
            "Dự án đã có Giấy phép xây dựng chưa?",
            "Dự án đã có quyết định giao đất của UBND tỉnh/thành phố chưa?",
        ]

    print(f"\n{BOLD}{'═'*60}")
    print("  DEBUG Q&A PIPELINE")
    print(f"{'═'*60}{RESET}")
    print(f"  Excel: {EXCEL_PATH}")
    print(f"  Project: {PROJECT_NAME}")

    # Bước 1
    rows, header_row = step1_read_excel(EXCEL_PATH)
    if rows is None:
        sys.exit(1)

    # Bước 2
    store = step2_load_store(EXCEL_PATH, PROJECT_NAME)
    if store is None:
        sys.exit(1)

    # Bước 3 — debug từng câu hỏi
    for query in test_queries:
        step3_debug_match(store, query, project_name=None, top_k=3)

    # Bước 4 — batch test
    step4_batch_test(store, project_name=None)

    print(f"\n{BOLD}{'═'*60}{RESET}")
    print(f"  {CYAN}Tip: Chạy với câu hỏi cụ thể:{RESET}")
    print(f"  python scripts/debug_qa.py \"Dự án này do CĐT tự thực hiện hay liên doanh/liên kết?\"")
    print(f"{BOLD}{'═'*60}{RESET}\n")


if __name__ == "__main__":
    main()
