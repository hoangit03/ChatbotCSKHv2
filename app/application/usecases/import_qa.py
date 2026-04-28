"""
app/application/usecases/import_qa.py

Use Case: Import bộ Q&A chuẩn từ file Excel vào QAVectorStore (Qdrant).

Excel format (tối thiểu 2 cột):
  | Question        | Answer       | Keywords (opt)      | DocGroup (opt) |
  |-----------------|--------------|---------------------|----------------|
  | Giá căn hộ 2PN? | Từ 3 tỷ VNĐ  | giá, 2pn, price     | price_list     |

Luồng xử lý:
  1. Đọc Excel → parse rows → tạo QAItem với deterministic ID
  2. Bỏ qua rows thiếu question/answer hoặc là dòng tiêu đề nhóm
  3. Gom toàn bộ items → 1 lần bulk_add (1 API call embedding cho cả batch)
  4. Qdrant upsert idempotent → import lại file cũ/file mới không duplicate

SRP: chỉ lo đọc Excel + normalize + add vào QAVectorStore.
OCP: format Excel thay đổi → chỉ sửa _parse_rows(), không sửa gì khác.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Optional

from app.agent.tools.qa_tool import QAItem, QAVectorStore
from app.shared.logging.logger import get_logger

log = get_logger(__name__)


# ── Request / Response DTOs ───────────────────────────────────────

@dataclass
class ImportQARequest:
    file_bytes: bytes
    file_name: str
    project_name: str
    sheet_name: Optional[str] = None     # None → sheet đầu tiên


@dataclass
class ImportQAResult:
    total_rows: int
    imported: int
    skipped: int
    errors: list[str] = field(default_factory=list)


# ── Use Case ──────────────────────────────────────────────────────

class ImportQAUseCase:
    """
    Đọc Excel → chuẩn hoá → bulk_add vào QAVectorStore.
    QAVectorStore được inject qua constructor (DIP).

    Async vì QAVectorStore.bulk_add() cần async embedding API call.
    Idempotent: QAItem.make_id() tạo deterministic UUID từ (project, question)
    → import cùng file nhiều lần, hoặc nhiều file có Q&A trùng → không duplicate.
    """

    def __init__(self, qa_store: QAVectorStore) -> None:
        self._store = qa_store

    async def execute(self, req: ImportQARequest) -> ImportQAResult:
        """
        Async: openpyxl parse sync → collect items → await bulk_add một lần.
        """
        log.info(
            "qa_import_start",
            project=req.project_name,
            file=req.file_name,
            sheet=req.sheet_name,
        )

        try:
            import openpyxl
        except ImportError:
            raise ImportError("pip install openpyxl")

        wb = openpyxl.load_workbook(io.BytesIO(req.file_bytes), data_only=True)

        # Chọn sheet
        if req.sheet_name and req.sheet_name in wb.sheetnames:
            ws = wb[req.sheet_name]
        else:
            ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return ImportQAResult(
                total_rows=0, imported=0, skipped=0,
                errors=["File Excel không có dữ liệu (cần ít nhất 1 dòng header + 1 dòng data)"],
            )

        header = [str(c or "").strip().lower() for c in rows[0]]
        col = _map_columns(header)

        if "question" not in col or "answer" not in col:
            return ImportQAResult(
                total_rows=0, imported=0, skipped=0,
                errors=["Excel thiếu cột 'Question' hoặc 'Answer'. "
                        "Kiểm tra tên cột ở dòng đầu tiên."],
            )

        items: list[QAItem] = []
        skipped = 0
        errors: list[str] = []
        data_rows = rows[1:]

        for row_idx, row in enumerate(data_rows, start=2):
            try:
                question = _cell(row, col["question"])
                answer   = _cell(row, col["answer"])

                if not question or not answer:
                    skipped += 1
                    continue

                # Bỏ qua dòng tiêu đề nhóm (vd: "1. Nhóm câu hỏi về pháp lý...")
                if re.match(r"^\d+[\.]\s+", question) and "?" not in question:
                    skipped += 1
                    continue

                keywords: list[str] = []
                if "keywords" in col:
                    raw_kw = _cell(row, col["keywords"])
                    keywords = [k.strip() for k in raw_kw.split(",") if k.strip()]

                doc_group: Optional[str] = None
                if "docgroup" in col:
                    doc_group = _cell(row, col["docgroup"]) or None

                item = QAItem(
                    id=QAItem.make_id(req.project_name, question),   # deterministic
                    project_name=req.project_name,
                    question=question,
                    answer=answer,
                    keywords=keywords,
                    doc_group=doc_group,
                )
                items.append(item)

            except Exception as e:
                errors.append(f"Dòng {row_idx}: {e}")

        # ── Bulk embed + upsert một lần — tối ưu API call ─────────
        imported = 0
        if items:
            try:
                imported = await self._store.bulk_add(items)
            except Exception as e:
                log.error("qa_import_bulk_add_failed", error=str(e))
                errors.append(f"Lỗi upsert Qdrant: {e}")

        log.info(
            "qa_import_done",
            project=req.project_name,
            total=len(data_rows),
            imported=imported,
            skipped=skipped,
            errors=len(errors),
        )

        return ImportQAResult(
            total_rows=len(data_rows),
            imported=imported,
            skipped=skipped,
            errors=errors,
        )


# ── Helpers ───────────────────────────────────────────────────────

def _cell(row: tuple, idx: int) -> str:
    try:
        return str(row[idx] or "").strip()
    except IndexError:
        return ""


# Alias linh hoạt cho tên cột — hỗ trợ cả tiếng Việt lẫn tiếng Anh
_COLUMN_ALIASES: dict[str, list[str]] = {
    "question": ["question", "câu hỏi", "cau hoi", "q", "hỏi"],
    "answer":   ["answer", "câu trả lời", "cau tra loi", "a", "trả lời", "tra loi"],
    "keywords": ["keywords", "từ khoá", "tu khoa", "tags", "từ khóa"],
    "docgroup": ["docgroup", "doc_group", "nhóm tài liệu", "group", "nhom"],
}


def _map_columns(header: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    for canonical, aliases in _COLUMN_ALIASES.items():
        for idx, col_name in enumerate(header):
            normalized = col_name.lower().strip().replace(" ", "")
            if col_name in aliases or normalized in [a.replace(" ", "") for a in aliases]:
                result[canonical] = idx
                break
    return result